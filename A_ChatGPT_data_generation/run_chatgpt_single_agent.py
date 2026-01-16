#!/usr/bin/env python3
"""Print first N rows of an XLSX file (sheet 1) to stdout in CSV style.

Usage:
  python3 run_chatgpt_single_agent.py [path_to_xlsx] [nrows] [file_id]

Defaults:
  path_to_xlsx: ./Questions_with_answer.xlsx
  nrows: 3 (number of rows to print as sample)
  file_id: None (used for output filename generation)

Requires: openpyxl
Install: pip3 install --user openpyxl
"""

import sys
import os
from pathlib import Path
import re
import time

# If you prefer to hard-code your OpenAI API key (not recommended),
# place it below as a string. Leave empty to use the environment variable
# `OPENAI_API_KEY` instead.
# Example (DO NOT commit real keys to version control):
# HARD_CODED_OPENAI_API_KEY = "sk-XXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXX"
HARD_CODED_OPENAI_API_KEY = None
SKIP_API_CALL = False  # Set to True to skip calling OpenAI API (for testing)
TEST_ITER_COUNT = None  # Number of rows to process in iterate_questions
CHATGPT_VERSION = "gpt-4"

try:
	from openpyxl import load_workbook
except Exception:
	print("Missing dependency: openpyxl. Install with: pip3 install --user openpyxl", file=sys.stderr)
	sys.exit(2)


def get_data(path, nrows=3): #nrows はサンプルとして表示する行数
	# Backwards-compatible: keep for callers that expect printing behavior
	data = load_sheet_to_list(path)
	count = 0
	for row in data[:nrows]:
		out = ",".join('"{}"'.format(c.replace('"', '""')) for c in row)
		count += 1
		print(out)
	# Return the loaded data so callers can reuse it
	return data


def load_sheet_to_list(path, max_rows=None):
	"""Load sheet1 into a list of rows (each row is list of str).

	- path: path to .xlsx
	- max_rows: if provided, limit reading to this many rows (None = all)
	Returns: list of rows (each row a list of strings)
	"""
	p = Path(path)
	if not p.exists():
		raise FileNotFoundError(f"File not found: {p}")
	wb = load_workbook(p, read_only=True, data_only=True)
	ws = wb.worksheets[0]
	rows = []
	for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
		row_vals = ["" if v is None else str(v) for v in row]
		rows.append(row_vals)
		if max_rows is not None and i >= max_rows:
			break
	return rows


def ask_openai(prompt, model=CHATGPT_VERSION, max_tokens=300):
	"""Send `prompt` to OpenAI Chat Completions (ChatGPT) and return assistant reply as string.

	Requires: environment variable `OPENAI_API_KEY` to be set.
	Uses `requests` (install with `pip3 install --user requests`).
	"""
	# Prefer a hard-coded key if provided (useful for quick local testing).
	# If HARD_CODED_OPENAI_API_KEY is empty, fall back to the environment variable.
	api_key = HARD_CODED_OPENAI_API_KEY or os.getenv("OPENAI_API_KEY")
	if not api_key:
		print("Missing OpenAI API key. Set HARD_CODED_OPENAI_API_KEY in the script or export OPENAI_API_KEY.", file=sys.stderr)
		print("  export OPENAI_API_KEY=\"sk-...\"", file=sys.stderr)
		return None
	if HARD_CODED_OPENAI_API_KEY:
		print("Warning: using hard-coded OpenAI API key from script (do not commit this key).", file=sys.stderr)
		pass

	try:
		import requests
	except Exception:
		print("Missing dependency: requests. Install with: pip3 install --user requests", file=sys.stderr)
		return None

	url = "https://api.openai.com/v1/chat/completions"
	headers = {
		"Authorization": f"Bearer {api_key}",
		"Content-Type": "application/json",
	}
	payload = {
		"model": model,
		"messages": [{"role": "user", "content": prompt}],
		"max_tokens": max_tokens,
	}

	try:
		r = requests.post(url, headers=headers, json=payload, timeout=30)
		r.raise_for_status()
		j = r.json()
		return j["choices"][0]["message"]["content"].strip()
	except Exception as e:
		print("Error calling OpenAI API:", e, file=sys.stderr)
		return None

def iterate_questions(data):
	for i, row in enumerate(data, start=0):
		if i == 0:
			continue  # Skip header row if present
		#print("Processing row", i, row)
		if TEST_ITER_COUNT is not None and i > TEST_ITER_COUNT:
			break
		if i > 2:
			break
		try:
			answer = None
			q_str = row[2] + ""
			if row[0] == "OP":
				q_str += "簡潔に答えだけを答えてください。"
			if row[0] == "4P":
				joined = "・".join(row[3:7])
				q_str += f"次の4つの中から選択してください。{joined}。"
			if row[0] == "2P":
				joined = "・".join(row[3:5])
				q_str += f"次の2つの中から選択してください。{joined}。"
			q_str += "またその答えの確信度を0-100%で答えてください。"
			q_str += "回答は 答え,確信度 という形式で答えてください。"
			print("Question:", i, q_str)
			if SKIP_API_CALL:
				print("Skipping OpenAI API call (testing mode).")
				answer = "noanswer, 0%"
				#continue
			else:
				answer = ask_openai(q_str)
				if answer is None:
					print("No answer received (see stderr for details).", file=sys.stderr)
					sys.exit(6)
				#print("Answer:", answer, "\n")
			print("Answer:", answer, "\n")

			# data に追記（I/J/K = index 8/9/10）
			row_idx = i
			while len(data[row_idx]) <= 10:
				data[row_idx].append("")
			
			# I列（index 8） = 生レスポンス
			data[row_idx][8] = answer

			# カンマで最初だけ分割して J/K 列へ
			parts = [p.strip() for p in answer.rsplit(",", 1)]
			# "," と "、" で分割
			# support both ASCII comma and Japanese comma, split at the last occurrence
			last_pos = max(answer.rfind(','), answer.rfind('、'))
			if last_pos == -1:
				parts = [answer.strip()]
			else:
				parts = [answer[:last_pos].strip(), answer[last_pos+1:].strip()]


			if len(parts) == 1:
				# カンマなし。答えだけ J 列へ
				data[row_idx][9] = parts[0]
				data[row_idx][10] = ""
			else:
				# カンマあり。答えを J 列、確信度を K 列へ
				data[row_idx][9] = parts[0]
				m = re.search(r"(\d+(?:\.\d+)?)", parts[-1])
				if m:
					num = float(m.group(1))
					if num.is_integer():
						num = int(num)
					data[row_idx][10] = num
				else:
					data[row_idx][10] = parts[-1]  # 数字が見つからなければそのまま

		except Exception as e:
			print("Error on row", i, e, file=sys.stderr)
			continue    
	
	return data

def write_data_to_excel(src_path, data, out_path=None, file_id=None):
	p = Path(src_path)

	results_dir = p.parent / "raw_responses"
	try:
		results_dir.mkdir(parents=True, exist_ok=True)
	except Exception as e:
		print("Could not create raw_responses directory:", e, file=sys.stderr)
	
	# ファイル名を決定（file_id があればゼロ埋め3桁）
	if file_id is not None:
		try:
			num = int(file_id)
			filename = f"ChatGPT-{num:03d}.xlsx"
		except Exception:
			filename = "ChatGPT-998.xlsx"
	else:
		filename = "ChatGPT-999.xlsx"


	# out_path 指定の解釈: None -> Results/filename、
	# ディレクトリ指定ならそのディレクトリ配下に filename を使用、
	# ファイルパス指定ならそのまま使用
	# out_path が未指定なら Results フォルダに <元名>_results.xlsx を作る
	if out_path is None:
		out_path = results_dir / filename
	else:
		out_path = Path(out_path)
		if out_path is None:
			out_path = results_dir / filename

	try:
		wb = load_workbook(p)
		ws = wb.worksheets[0]
	except Exception:
		from openpyxl import Workbook
		wb = Workbook()
		ws = wb.active

	for i, row in enumerate(data, start=1):
		for j, val in enumerate(row, start=1):
			ws.cell(row=i, column=j, value=val)

	try:
		wb.save(out_path)
		return str(out_path)	
	except Exception as e:
		print("Error saving Excel file:", e, file=sys.stderr)
		return None


if __name__ == "__main__":
	path = sys.argv[1] if len(sys.argv) > 1 else "Questions_with_answer.xlsx"
	# 2 = nrows: development-only preview; keep default (3) in batch runs
	try:
		nrows = int(sys.argv[2]) if len(sys.argv) > 2 else 3
	except ValueError:
		print("nrows must be an integer", file=sys.stderr)
		sys.exit(4)

    # File number (ex: 3 -> ChatGPT-003.xlsx）
	try:
		file_id = int(sys.argv[3]) if len(sys.argv) > 3 else None
	except ValueError:
		print("file id must be an integer", file=sys.stderr)
		sys.exit(4)	

	# Capture returned data if the caller (or another part of the program)
	# wants to use it after printing.
	data = get_data(path, nrows)
	print("\n--- Now iterating questions and asking OpenAI ---\n", file=sys.stderr)
	data = iterate_questions(data)
	out_file = write_data_to_excel(path, data, file_id=file_id)
	print("Saved results to", out_file)
	#time.sleep(10)   # wait for 10 seconds to avoid rate limits




